"""
===================================================================
  GEO miRNA Cross-Platform Integration Pipeline
  For PDAC (Pancreatic Ductal Adenocarcinoma) Research

  Features:
    - Universal GEO Series Matrix reader (.txt and .xlsx)
    - Automatic platform detection and scale inference
    - Probe ID harmonization across platforms
    - Z-score normalization per dataset
    - Cross-dataset merge by common miRNAs
    - ComBat batch correction (preserving biological signal)
    - PurityB / PurityD validation
    - PCA and optional UMAP visualization

  Supported platforms:
    Affymetrix (GPL19117, GPL18402, etc.)
    3D-Gene / Toray (GPL18941, GPL21263)
    Agilent, Illumina, and others

  Usage:
    python geo_mirna_pipeline.py GSE85589_series_matrix.txt \\
        GSE59856_series_matrix.txt --output-root ./out

  Author: TCC Pipeline (evolved from app.py)
===================================================================
"""

import re
import os
import sys
import io
import logging
import argparse
import warnings
import subprocess
from pathlib import Path
from typing import Dict, List, Optional, Tuple
from collections import Counter

# Fix Windows console encoding for emoji/unicode
if sys.stdout.encoding and sys.stdout.encoding.lower() not in ("utf-8", "utf8"):
    try:
        sys.stdout = io.TextIOWrapper(
            sys.stdout.buffer, encoding="utf-8", errors="replace",
        )
        sys.stderr = io.TextIOWrapper(
            sys.stderr.buffer, encoding="utf-8", errors="replace",
        )
    except Exception:
        pass

import numpy as np
import pandas as pd
from sklearn.cluster import KMeans
from sklearn.decomposition import PCA
from sklearn.preprocessing import LabelEncoder

import matplotlib
matplotlib.use("Agg")  # non-interactive backend for plot generation
import matplotlib.pyplot as plt

# ── Optional dependency checks ──────────────────────────────────────
try:
    import chardet
    HAS_CHARDET = True
except ImportError:
    HAS_CHARDET = False

try:
    import seaborn as sns
    HAS_SEABORN = True
except ImportError:
    HAS_SEABORN = False

try:
    import umap as umap_lib
    HAS_UMAP = True
except ImportError:
    HAS_UMAP = False


# ── Logging ─────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger("geo_pipeline")


# =====================================================================
# Constants
# =====================================================================

KNOWN_PLATFORMS: Dict[str, str] = {
    "GPL19117": "Affymetrix miRNA 4.0 (multispecies)",
    "GPL18941": "3D-Gene Human miRNA (Toray)",
    "GPL21263": "3D-Gene Human miRNA V21 (Toray)",
    "GPL18402": "Affymetrix miRNA 4.0",
    "GPL16384": "Affymetrix miRNA 3.0",
    "GPL8786":  "Affymetrix miRNA 1.0",
    "GPL10850": "Agilent Human miRNA V2",
    "GPL18058": "Agilent Human miRNA V3",
    "GPL11487": "Agilent Human miRNA V4",
    "GPL7731":  "Agilent Human miRNA V1",
    "GPL8179":  "Illumina Human v2",
    "GPL6480":  "Agilent Whole Human Genome",
    "GPL570":   "Affymetrix HG-U133 Plus 2.0",
    "GPL96":    "Affymetrix HG-U133A",
}


# =====================================================================
# 1. CLI (argparse)
# =====================================================================

def build_cli() -> argparse.Namespace:
    """Build and parse command-line arguments."""
    parser = argparse.ArgumentParser(
        description=(
            "GEO miRNA Cross-Platform Integration Pipeline "
            "for PDAC (Pancreatic Cancer) Research"
        ),
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Interactive mode (opens file picker):
  python geo_mirna_pipeline.py

  # Process two datasets, merge, apply ComBat
  python geo_mirna_pipeline.py GSE85589_series_matrix.txt \\
      GSE59856_series_matrix.txt --output-root ./out

  # Non-interactive mode with condition filter
  python geo_mirna_pipeline.py GSE85589.txt GSE59856.txt \\
      --no-interactive \\
      --condition-filter "pancreatic cancer" "healthy control" \\
      --class-map "pancreatic cancer=PDAC" "healthy control=Control"

  # Single dataset, no merge
  python geo_mirna_pipeline.py GSE85589_series_matrix.txt \\
      --output-root ./results
        """,
    )
    parser.add_argument(
        "files", nargs="*",
        help="GEO Series Matrix files (.txt or .xlsx). If omitted, a file picker dialog opens.",
    )
    parser.add_argument(
        "--output-root", default=".",
        help="Root output directory (default: current directory)",
    )
    parser.add_argument(
        "--no-interactive", action="store_true",
        help="Skip interactive prompts; use all samples if no filter given",
    )
    parser.add_argument(
        "--condition-filter", nargs="*", default=None,
        help=(
            "Conditions to keep (substring match, applied to all datasets). "
            "E.g.: --condition-filter \"pancreatic cancer\" \"healthy control\""
        ),
    )
    parser.add_argument(
        "--class-map", nargs="*", default=None,
        help=(
            "Map condition names to class labels. "
            "E.g.: --class-map \"pancreatic cancer=PDAC\" \"healthy control=Control\""
        ),
    )
    parser.add_argument(
        "--no-combat", action="store_true",
        help="Skip ComBat batch correction",
    )
    parser.add_argument(
        "--zscore-only", action="store_true",
        help="Only do z-score merge, skip ComBat",
    )
    parser.add_argument(
        "--no-plots", action="store_true",
        help="Skip plot generation",
    )
    return parser.parse_args()


def interactive_file_picker() -> List[str]:
    """
    Open a native file picker dialog to select GEO Series Matrix files.
    Uses tkinter (bundled with Python) so no extra install is needed.
    Files can be selected from ANY folder on the computer.

    Supports multiple rounds of selection so the user can pick files
    from different folders (e.g., GSE85589 from one folder and
    GSE59856 from another).

    Returns:
        List of absolute file paths selected by the user.
    """
    try:
        import tkinter as tk
        from tkinter import filedialog
    except ImportError:
        log.error(
            "tkinter not available. Please provide files via command line:\n"
            "  python geo_mirna_pipeline.py file1.txt file2.txt"
        )
        sys.exit(1)

    all_files: List[str] = []

    while True:
        root = tk.Tk()
        root.withdraw()
        root.attributes("-topmost", True)

        if not all_files:
            print("\n  Abrindo seletor de arquivos...")
            print("  (Selecione um ou mais arquivos GEO Series Matrix)")
            print("  Dica: Segure Ctrl para selecionar varios arquivos!\n")
        else:
            print(f"\n  Abrindo seletor para adicionar mais arquivos...")
            print(f"  ({len(all_files)} arquivo(s) ja selecionado(s))\n")

        file_paths = filedialog.askopenfilenames(
            title=f"Selecione arquivos GEO Series Matrix ({len(all_files)} ja selecionados)",
            filetypes=[
                ("GEO Series Matrix", "*.txt *.xlsx *.xls"),
                ("Arquivos de texto", "*.txt"),
                ("Planilhas Excel", "*.xlsx *.xls"),
                ("Todos os arquivos", "*.*"),
            ],
        )

        root.destroy()

        if file_paths:
            for f in file_paths:
                if f not in all_files:
                    all_files.append(f)
                    print(f"  + {Path(f).name}")

        if not all_files:
            return []

        print(f"\n  Total: {len(all_files)} arquivo(s) selecionado(s)")

        # Ask if user wants to add more files (from another folder)
        try:
            more = input(
                "  Deseja adicionar mais arquivos de outra pasta? (s/N): "
            ).strip().lower()
            if more not in ("s", "sim", "y", "yes"):
                break
        except (EOFError, KeyboardInterrupt):
            break

    return all_files


def interactive_output_picker() -> str:
    """
    Ask the user where to save the output, with a friendly console menu.

    Returns:
        Path string for the output directory.
    """
    print("\n" + "-" * 50)
    print("  Onde salvar os resultados?")
    print("-" * 50)
    print("  [1] Pasta atual (.)")
    print("  [2] Escolher pasta via seletor")
    print("  [3] Digitar caminho manualmente")
    print("-" * 50)

    while True:
        try:
            choice = input("\n  Escolha (1/2/3) [padrao: 1]: ").strip()
            if choice in ("", "1"):
                return "."
            elif choice == "2":
                try:
                    import tkinter as tk
                    from tkinter import filedialog

                    root = tk.Tk()
                    root.withdraw()
                    root.attributes("-topmost", True)

                    folder = filedialog.askdirectory(
                        title="Selecione a pasta para salvar os resultados",
                    )
                    root.destroy()

                    if folder:
                        return folder
                    else:
                        print("  Nenhuma pasta selecionada. Usando pasta atual.")
                        return "."
                except ImportError:
                    print("  tkinter nao disponivel. Digite o caminho:")
                    path = input("  Caminho: ").strip()
                    return path if path else "."
            elif choice == "3":
                path = input("  Caminho da pasta: ").strip()
                return path if path else "."
            else:
                print("  Opcao invalida. Digite 1, 2 ou 3.")
        except (EOFError, KeyboardInterrupt):
            print("\n  Usando pasta atual.")
            return "."


# =====================================================================
# 2. Encoding Detection
# =====================================================================

def detect_encoding(path: str) -> str:
    """Detect file encoding using chardet, falling back to latin-1."""
    if not HAS_CHARDET:
        log.warning("chardet not installed; assuming latin-1 encoding")
        return "latin-1"
    with open(path, "rb") as fh:
        raw = fh.read(2_000_000)
    result = chardet.detect(raw)
    enc = result.get("encoding") or "latin-1"
    conf = result.get("confidence", "?")
    log.info(f"Encoding detected: {enc} (confidence: {conf})")
    return enc


# =====================================================================
# 3. Metadata Parsing
# =====================================================================

def parse_series_metadata_tabular(path: str) -> pd.DataFrame:
    """
    Parse sample-level metadata from a GEO Series Matrix file.

    Handles both .txt (tab-delimited) and .xlsx formats.
    Fixes the duplicate-key bug where multiple !Sample_characteristics_ch1
    lines would overwrite each other.
    """
    ext = Path(path).suffix.lower()
    if ext in (".xlsx", ".xls"):
        return _parse_metadata_xlsx(path)
    return _parse_metadata_txt(path)


def _make_keys_unique(data: List[Tuple[str, list]]) -> List[Tuple[str, list]]:
    """
    Ensure all metadata keys are unique by appending _2, _3, …
    to duplicates.  Fixes the GEO bug where multiple
    !Sample_characteristics_ch1 lines exist.
    """
    seen: Dict[str, int] = {}
    unique_data: List[Tuple[str, list]] = []
    for key, vals in data:
        if key in seen:
            seen[key] += 1
            unique_key = f"{key}_{seen[key]}"
        else:
            seen[key] = 1
            unique_key = key
        unique_data.append((unique_key, vals))
    return unique_data


def _build_metadata_df(data: List[Tuple[str, list]]) -> pd.DataFrame:
    """Convert list of (key, values) pairs into a sample × field DataFrame."""
    if not data:
        log.warning("No metadata key/value pairs found")
        return pd.DataFrame()

    data = _make_keys_unique(data)
    max_cols = max(len(v) for _, v in data)

    df = pd.DataFrame(
        {k: v + [""] * (max_cols - len(v)) for k, v in data}
    ).T
    df.reset_index(inplace=True)
    df.rename(columns={"index": "Field"}, inplace=True)
    df = df.set_index("Field").T.reset_index(drop=True)

    log.info(f"Metadata parsed: {df.shape[0]} samples × {df.shape[1]} fields")
    return df


def _parse_metadata_txt(path: str) -> pd.DataFrame:
    """Parse metadata from a GEO Series Matrix .txt file."""
    enc = detect_encoding(path)

    meta_lines: List[str] = []
    with open(path, "r", encoding=enc, errors="ignore") as fh:
        for line in fh:
            if line.lower().startswith("!series_matrix_table_begin"):
                break
            if line.strip().startswith("!"):
                meta_lines.append(line.strip())

    data: List[Tuple[str, list]] = []
    for raw_line in meta_lines:
        parts = raw_line.split("\t")
        if len(parts) > 1:
            key = parts[0].lstrip("!").strip()
            vals = [v.strip().strip('"') for v in parts[1:] if v.strip()]
            data.append((key, vals))

    return _build_metadata_df(data)


def _parse_metadata_xlsx(path: str) -> pd.DataFrame:
    """Parse metadata from a GEO Series Matrix .xlsx file."""
    try:
        import openpyxl
    except ImportError:
        log.error("openpyxl required for .xlsx files: pip install openpyxl")
        return pd.DataFrame()

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    data: List[Tuple[str, list]] = []
    for row in ws.iter_rows(values_only=True):
        first_cell = str(row[0]) if row[0] else ""
        if first_cell.lower().startswith("!series_matrix_table_begin"):
            break
        if first_cell.startswith("!"):
            key = first_cell.lstrip("!").strip()
            vals = [
                str(v).strip().strip('"')
                for v in row[1:]
                if v is not None and str(v).strip()
            ]
            data.append((key, vals))

    wb.close()
    return _build_metadata_df(data)


# =====================================================================
# 4. Platform Detection
# =====================================================================

def detect_platform(meta_df: pd.DataFrame) -> Tuple[str, str]:
    """
    Detect the microarray platform from metadata.

    Returns:
        (platform_id, platform_name)
    """
    # Search dedicated platform columns
    platform_cols = [
        c for c in meta_df.columns
        if any(k in c.lower() for k in ("platform_id",))
    ]

    platform_id = "Unknown"
    for col in platform_cols:
        vals = meta_df[col].astype(str).unique()
        vals = [v for v in vals if v and v not in ("", "nan")]
        if vals:
            platform_id = vals[0]
            break

    # Fallback: scan all columns for GPL pattern
    if platform_id == "Unknown":
        for col in meta_df.columns:
            for val in meta_df[col].astype(str):
                m = re.search(r"(GPL\d+)", str(val))
                if m:
                    platform_id = m.group(1)
                    break
            if platform_id != "Unknown":
                break

    platform_name = KNOWN_PLATFORMS.get(platform_id, platform_id)
    log.info(f"🔬 Platform: {platform_id} ({platform_name})")
    return platform_id, platform_name


# =====================================================================
# 5. Scale Inference
# =====================================================================

def infer_scale(
    meta_df: pd.DataFrame,
    platform_id: str,
    expr_num: pd.DataFrame,
) -> str:
    """
    Infer the data scale from metadata keywords, platform rules,
    and value distribution.

    Returns one of:
        "already_log2"       – RMA or explicit log2 data
        "already_processed"  – normalized but scale ambiguous
        "needs_log2"         – raw intensities requiring log2(x+1)
        "unknown"            – cannot determine
    """
    # ── 1. Extract data_processing text ──
    processing_text = ""
    for col in meta_df.columns:
        if "data_processing" in col.lower():
            vals = meta_df[col].dropna().unique()
            processing_text = " ".join(str(v).lower() for v in vals)
            break

    has_log2   = any(k in processing_text for k in ("log2", "log 2", "log-transformed"))
    # Use word-boundary match for RMA etc. to avoid matching 'normalization'
    has_rma    = bool(re.search(r"\brma\b", processing_text)) or \
                 bool(re.search(r"\bmas5\b", processing_text)) or \
                 bool(re.search(r"\bgcrma\b", processing_text))
    has_qnorm  = "quantile" in processing_text
    has_norm   = "normalized" in processing_text or "normalisation" in processing_text

    # ── 2. Value distribution ──
    gsm_cols = [c for c in expr_num.columns if c.startswith("GSM")]
    if gsm_cols:
        flat = expr_num[gsm_cols].values.flatten()
        flat = flat[np.isfinite(flat)]
        if len(flat) > 0:
            v_min, v_med, v_max = (
                np.nanmin(flat), np.nanmedian(flat), np.nanmax(flat),
            )
        else:
            v_min = v_med = v_max = 0.0
    else:
        v_min = v_med = v_max = 0.0

    log.info(
        f"Value stats: min={v_min:.3f}, median={v_med:.3f}, max={v_max:.3f}"
    )

    # ── 3. Platform rules ──
    plat_name = KNOWN_PLATFORMS.get(platform_id, "").lower()
    is_affy   = "affymetrix" in plat_name or "affy" in platform_id.lower()
    is_3dgene = "3d-gene" in plat_name or "toray" in plat_name

    # ── Decision logic ──
    if has_rma:
        log.info("Scale: RMA keyword found → already_log2")
        return "already_log2"

    if has_log2:
        log.info("Scale: log2 keyword found → already_log2")
        return "already_log2"

    # Low value range + evidence of processing
    if v_max < 25 and v_min >= -10 and v_med > 0:
        if has_qnorm or has_norm:
            log.info("Scale: low range + normalization keyword → already_log2")
            return "already_log2"
        if is_affy or is_3dgene:
            log.info(f"Scale: low range + known platform → already_log2")
            return "already_log2"
        log.info("Scale: low range → already_processed (verify manually)")
        return "already_processed"

    # High values → raw intensities
    if v_max > 1000:
        log.info("Scale: high values → needs_log2")
        return "needs_log2"

    log.warning("Scale: could not determine → unknown")
    return "unknown"


# =====================================================================
# 6. Numeric Parsing
# =====================================================================

def smart_float(x) -> float:
    """
    Convert expression value to float with robust handling of:

    - Multi-dot locale formatting: "1.129.491.157" → 1.129491157
      (common when .xlsx files are exported with PT-BR locale where
       dots are thousand separators; the original value is 1.129491157)
    - Standard decimals: "2.612" → 2.612
    - Missing / NA markers → NaN

    In genuine GEO Series Matrix .txt files, the dot is ALWAYS the
    decimal separator and there are NO thousand separators.
    """
    s = str(x).replace('"', "").strip()

    if not s or s.lower() in ("na", "nan", "null", "none", "--", "n/a"):
        return np.nan

    # Multi-dot fix: keep first dot as decimal, remove the rest
    dot_count = s.count(".")
    if dot_count > 1:
        parts = s.split(".")
        s = parts[0] + "." + "".join(parts[1:])

    # Remove commas (never decimal in GEO)
    s = s.replace(",", "")

    try:
        return float(s)
    except ValueError:
        return np.nan


# =====================================================================
# 7. Expression Reading
# =====================================================================

def read_expression(path: str) -> Tuple[pd.DataFrame, pd.DataFrame, List[str]]:
    """
    Read expression data from a GEO Series Matrix file.

    Auto-detects format (.txt vs .xlsx) by extension.

    Returns:
        expr_text – DataFrame with original text values preserved
        expr_num  – DataFrame with parsed numeric values
        gsm_cols  – list of GSM sample column names
    """
    ext = Path(path).suffix.lower()
    if ext in (".xlsx", ".xls"):
        return _read_expression_xlsx(path)
    return _read_expression_txt(path)


def _read_expression_txt(path: str) -> Tuple[pd.DataFrame, pd.DataFrame, List[str]]:
    """Read expression table from a tab-delimited Series Matrix .txt."""
    enc = detect_encoding(path)

    # Find the header line containing ID_REF
    header_idx = None
    with open(path, "r", encoding=enc, errors="ignore") as fh:
        for i, line in enumerate(fh):
            if "ID_REF" in line:
                header_idx = i
                break

    if header_idx is None:
        raise ValueError("❌ Header containing 'ID_REF' not found in file")

    df = pd.read_csv(
        path, sep="\t", skiprows=header_idx, header=0,
        dtype=str, encoding=enc, engine="python",
    )

    # Clean up
    df = df[df.iloc[:, 0].notna()]
    df = df[~df.iloc[:, 0].astype(str).str.startswith("!")]

    first_col = df.columns[0]
    df.rename(columns={first_col: "Probe_ID"}, inplace=True)
    df["Probe_ID"] = df["Probe_ID"].astype(str).str.strip().str.strip('"')

    # Identify GSM columns and strip quotes from names
    gsm_cols_raw = [c for c in df.columns if re.match(r'^"?GSM\d+', str(c))]
    rename_map = {c: c.strip('"') for c in gsm_cols_raw}
    df.rename(columns=rename_map, inplace=True)
    gsm_cols = [rename_map[c] for c in gsm_cols_raw]

    expr_text = df[["Probe_ID"] + gsm_cols].copy()
    expr_num = expr_text.copy()

    for c in gsm_cols:
        expr_num[c] = expr_num[c].apply(smart_float)

    log.info(f"Expression loaded: {expr_num.shape[0]} probes × {len(gsm_cols)} samples")
    return expr_text, expr_num, gsm_cols


def _read_expression_xlsx(path: str) -> Tuple[pd.DataFrame, pd.DataFrame, List[str]]:
    """Read expression table from an .xlsx Series Matrix file."""
    try:
        import openpyxl
    except ImportError:
        raise ImportError("openpyxl required for .xlsx: pip install openpyxl")

    log.warning(
        "⚠️  Reading from .xlsx — numeric values may have locale issues. "
        "Prefer the original .txt Series Matrix from GEO."
    )

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    header_row = None
    rows_data: List[tuple] = []
    in_table = False

    for row in ws.iter_rows(values_only=True):
        first = str(row[0]) if row[0] else ""
        if "series_matrix_table_begin" in first.lower():
            in_table = True
            continue
        if "series_matrix_table_end" in first.lower():
            break
        if in_table:
            if header_row is None:
                header_row = [
                    str(v).strip().strip('"') if v else ""
                    for v in row
                ]
            else:
                rows_data.append(row)

    wb.close()

    if header_row is None:
        raise ValueError("Expression table not found in xlsx")

    df = pd.DataFrame(rows_data, columns=header_row)
    df.rename(columns={df.columns[0]: "Probe_ID"}, inplace=True)
    df["Probe_ID"] = df["Probe_ID"].astype(str).str.strip().str.strip('"')

    gsm_cols = [c for c in df.columns if re.match(r"^GSM\d+", str(c))]

    expr_text = df[["Probe_ID"] + gsm_cols].copy()
    expr_num = expr_text.copy()

    for c in gsm_cols:
        expr_num[c] = expr_num[c].apply(smart_float)

    # Sanity check: warn if values are suspiciously large
    sample_vals = expr_num[gsm_cols].values.flatten()
    sample_vals = sample_vals[np.isfinite(sample_vals)]
    if len(sample_vals) > 0 and np.nanmax(sample_vals) > 1_000_000:
        log.warning(
            "⚠️  Extremely large values detected (max=%.2e). "
            "This usually indicates locale-corrupted .xlsx data. "
            "Please use the original .txt Series Matrix file.",
            np.nanmax(sample_vals),
        )

    log.info(
        f"Expression (xlsx): {expr_num.shape[0]} probes × {len(gsm_cols)} samples"
    )
    return expr_text, expr_num, gsm_cols


# =====================================================================
# 8. Condition Extraction & Filtering
# =====================================================================

def normalize_condition(value: str) -> str:
    """
    Remove individual sample IDs from the end of a condition string.

    Examples:
        "pancreatic cancer P75"   → "pancreatic cancer"
        "healthy control E001"    → "healthy control"
        "biliary tract cancer B101" → "biliary tract cancer"
    """
    normalized = re.sub(r"\s+[A-Z]{0,2}\d{1,4}\s*$", "", value.strip())
    return normalized.strip()


def _extract_sample_condition(
    meta_row: pd.Series,
    all_columns: List[str],
) -> str:
    """
    Extract the most informative disease/condition label from a metadata row.

    Priority:
        1. Characteristics column containing "disease state: …"
        2. Source name with parenthesised condition
        3. Title (normalised)
    """
    # Priority 1: characteristics with "disease" keyword
    for col in all_columns:
        if "characteristics" not in col.lower():
            continue
        val = str(meta_row.get(col, ""))
        if "disease" in val.lower() and ":" in val:
            # "disease state: pancreatic cancer" → "pancreatic cancer"
            return re.sub(r"^[^:]+:\s*", "", val).strip().strip('"')

    # Priority 2: source_name (extract from parentheses)
    for col in all_columns:
        if "source_name" not in col.lower():
            continue
        val = str(meta_row.get(col, "")).strip().strip('"')
        m = re.search(r"\(([^)]+)\)", val)
        if m:
            return m.group(1).strip()
        # If source_name itself is descriptive (not just "Serum")
        if val and val.lower() not in ("", "nan", "none", "serum", "plasma", "blood"):
            return val

    # Priority 3: title
    for col in all_columns:
        if "title" not in col.lower():
            continue
        val = str(meta_row.get(col, "")).strip().strip('"')
        if val and val.lower() not in ("", "nan", "none"):
            return normalize_condition(val)

    return ""


def extract_conditions(
    meta_df: pd.DataFrame,
) -> Tuple[List[Tuple[str, int]], List[str]]:
    """
    Extract unique conditions from metadata, grouped by disease/condition.

    Uses the per-sample smart extraction logic (_extract_sample_condition)
    to correctly group samples by their actual disease category instead
    of showing individual patient entries.

    Returns:
        grouped    – list of (condition_name, sample_count) sorted by count desc
        cond_cols  – column names that contain condition information
    """
    condition_keywords = [
        "source_name", "characteristics", "title",
        "description", "disease", "tissue", "cell_type", "treatment",
    ]

    cond_cols: List[str] = []
    for col in meta_df.columns:
        col_lower = col.lower()
        if any(k in col_lower for k in condition_keywords):
            cond_cols.append(col)

    all_columns = list(meta_df.columns)
    conditions: List[str] = []

    for _, row in meta_df.iterrows():
        cond = _extract_sample_condition(row, all_columns)
        normalized = normalize_condition(cond) if cond else ""
        if normalized:
            conditions.append(normalized)

    counts: Counter = Counter(conditions)
    # Sort by count descending, then alphabetically
    grouped = sorted(counts.items(), key=lambda x: (-x[1], x[0].lower()))
    return grouped, cond_cols


def select_conditions_cli(
    grouped_conditions: List[Tuple[str, int]],
    condition_filter: Optional[List[str]] = None,
    no_interactive: bool = False,
) -> Optional[List[str]]:
    """
    Select which conditions to keep.

    Priority:
        1. --condition-filter from CLI (if provided)
        2. Interactive console prompt (if allowed)
        3. None (keep all samples)
    """
    if condition_filter:
        log.info(f"Using CLI condition filter: {condition_filter}")
        return condition_filter

    if no_interactive:
        log.info("Non-interactive mode: using ALL samples (no filter)")
        return None

    if not grouped_conditions:
        log.warning("No conditions found in metadata")
        return None

    print("\n" + "=" * 60)
    print("  Conditions / categories found in dataset:")
    print("=" * 60)
    for i, (cond, cnt) in enumerate(grouped_conditions, 1):
        print(f"  [{i}] {cond}  ({cnt} amostras)")
    print(f"  [0] Use ALL samples (no filter)")
    print("=" * 60)

    while True:
        try:
            raw = input(
                "\n🎯 Enter condition numbers separated by comma "
                "(e.g. 1,3) or 0 for all: "
            ).strip()

            if raw == "0":
                return None

            indices = [int(x.strip()) for x in raw.split(",")]
            selected: List[str] = []
            for idx in indices:
                if 1 <= idx <= len(grouped_conditions):
                    selected.append(grouped_conditions[idx - 1][0])
                else:
                    print(f"  ❌ Invalid number: {idx}")

            if selected:
                for s in selected:
                    cnt = dict(grouped_conditions).get(s, 0)
                    print(f"  ✅ '{s}' ({cnt} amostras)")
                return selected

        except ValueError:
            print("  ❌ Enter numbers only.")
        except (EOFError, KeyboardInterrupt):
            print("\n  ⚠️ Using all samples (no filter).")
            return None


def filter_samples_by_conditions(
    meta_df: pd.DataFrame,
    conditions: Optional[List[str]],
    condition_cols: List[str],
) -> pd.DataFrame:
    """
    Filter metadata rows matching any of the selected conditions
    (case-insensitive substring match).
    """
    if conditions is None:
        log.info(f"Using ALL {meta_df.shape[0]} samples (no filter)")
        return meta_df

    mask = pd.Series(False, index=meta_df.index)

    for cond in conditions:
        cond_lower = cond.lower()
        for col in condition_cols:
            col_mask = (
                meta_df[col]
                .astype(str)
                .str.lower()
                .str.contains(re.escape(cond_lower), na=False)
            )
            mask = mask | col_mask

    # Fallback: search ALL columns
    if mask.sum() == 0:
        log.warning("Primary columns had no matches; searching all columns")
        df_str = meta_df.astype(str).apply(lambda x: x.str.lower())
        for cond in conditions:
            cond_lower = cond.lower()
            mask = mask | df_str.apply(
                lambda row: cond_lower in " ".join(row.values), axis=1
            )

    filtered = meta_df.loc[mask].copy()
    log.info(f"Filtered: {filtered.shape[0]} samples for conditions {conditions}")
    return filtered


# =====================================================================
# 9. Probe ID Harmonization
# =====================================================================

def canonicalize_probe_id(probe_id: str, platform_id: str) -> Tuple[str, bool]:
    """
    Canonicalize a probe ID based on platform-specific rules.

    Args:
        probe_id:    original Probe_ID string
        platform_id: GPL accession

    Returns:
        (canonical_id, is_ambiguous)

    Rules:
        Affymetrix: remove trailing '_st'
        3D-Gene:    split comma-separated MIMATs; flag if >1
        Other:      pass through, strip whitespace
    """
    pid = probe_id.strip()

    # Affymetrix family
    affy_platforms = {"GPL19117", "GPL18402", "GPL16384", "GPL8786"}
    if platform_id in affy_platforms:
        canonical = re.sub(r"_st$", "", pid, flags=re.IGNORECASE)
        return canonical, False

    # 3D-Gene / Toray family
    toray_platforms = {"GPL18941", "GPL21263"}
    if platform_id in toray_platforms:
        if "," in pid:
            # Multiple MIMATs → ambiguous probe
            parts = [p.strip() for p in pid.split(",")]
            return parts[0], True  # use first, flag ambiguous
        return pid, False

    # Default
    return pid, False


def build_feature_map(
    expr_df: pd.DataFrame,
    platform_id: str,
) -> pd.DataFrame:
    """
    Build a probe-to-canonical-ID mapping with ambiguity flags.

    Returns DataFrame:
        Probe_ID  |  Probe_ID_Canonical  |  Probe_ID_Ambiguous
    """
    records = []
    for pid in expr_df["Probe_ID"].unique():
        canonical, ambiguous = canonicalize_probe_id(pid, platform_id)
        records.append({
            "Probe_ID": pid,
            "Probe_ID_Canonical": canonical,
            "Probe_ID_Ambiguous": ambiguous,
        })

    fmap = pd.DataFrame(records)
    n_ambig = fmap["Probe_ID_Ambiguous"].sum()
    log.info(f"Feature map: {len(fmap)} probes, {n_ambig} ambiguous")
    return fmap


# =====================================================================
# 10. Sample Annotation
# =====================================================================



def build_sample_annotation(
    meta_df: pd.DataFrame,
    gsm_cols: List[str],
    dataset_id: str,
    platform_id: str,
    platform_name: str,
    class_map: Optional[Dict[str, str]] = None,
) -> pd.DataFrame:
    """
    Build a standardised sample_annotation DataFrame.

    Columns:
        sample_id, dataset_id, batch, platform_id, platform_name,
        condition_raw, condition_normalized, class_label
    """
    # Find the GSM accession column (prioritise Sample_ over Series_)
    gsm_col: Optional[str] = None
    for col in meta_df.columns:
        if "sample_geo_accession" in col.lower().replace(" ", "_"):
            gsm_col = col
            break
    if gsm_col is None:
        for col in meta_df.columns:
            if "geo_accession" in col.lower():
                # Verify it actually contains GSM IDs, not GSE IDs
                if meta_df[col].astype(str).str.match(r"^GSM\d+").any():
                    gsm_col = col
                    break
    if gsm_col is None:
        for col in meta_df.columns:
            if meta_df[col].astype(str).str.match(r"^GSM\d+").any():
                gsm_col = col
                break

    all_columns = list(meta_df.columns)
    records: List[dict] = []

    for gsm_id in gsm_cols:
        condition_raw = ""

        # Look up sample row
        if gsm_col and gsm_id in meta_df[gsm_col].values:
            row = meta_df[meta_df[gsm_col] == gsm_id].iloc[0]
            condition_raw = _extract_sample_condition(row, all_columns)

        condition_normalized = normalize_condition(condition_raw)

        # Map to class label
        class_label = condition_normalized
        if class_map:
            for pattern, label in class_map.items():
                if pattern.lower() in condition_normalized.lower():
                    class_label = label
                    break

        records.append({
            "sample_id": gsm_id,
            "dataset_id": dataset_id,
            "batch": dataset_id,
            "platform_id": platform_id,
            "platform_name": platform_name,
            "condition_raw": condition_raw,
            "condition_normalized": condition_normalized,
            "class_label": class_label,
        })

    annot = pd.DataFrame(records)
    class_counts = annot["class_label"].value_counts().to_dict()
    log.info(f"Sample annotation: {len(annot)} samples, classes: {class_counts}")
    return annot


# =====================================================================
# 11. Z-Score Normalization
# =====================================================================

def zscore_by_probe(expr_df: pd.DataFrame) -> pd.DataFrame:
    """
    Z-score normalize each probe (row) across samples within a dataset.

        z = (x − mean_probe) / std_probe

    Probes with zero variance get z = 0.
    """
    result = expr_df.copy()
    gsm_cols = [c for c in result.columns if c.startswith("GSM")]

    data = result[gsm_cols].values.astype(float)

    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        means = np.nanmean(data, axis=1, keepdims=True)
        stds = np.nanstd(data, axis=1, keepdims=True, ddof=0)
        stds[stds == 0] = 1.0  # avoid division by zero
        z_data = (data - means) / stds

    result[gsm_cols] = z_data
    log.info(f"Z-score: {data.shape[0]} probes × {len(gsm_cols)} samples")
    return result


# =====================================================================
# 12. Per-Dataset Processing Pipeline
# =====================================================================

def extract_dataset_id(path: str) -> str:
    """Extract the GSE accession from a filename."""
    m = re.search(r"(GSE\d+)", Path(path).stem, re.IGNORECASE)
    return m.group(1) if m else Path(path).stem


def process_single_dataset(
    path: str,
    output_root: Path,
    no_interactive: bool = False,
    condition_filter: Optional[List[str]] = None,
    class_map: Optional[Dict[str, str]] = None,
) -> Optional[Path]:
    """
    Run the full per-dataset processing pipeline on a single
    GEO Series Matrix file.

    Outputs (in out_<GSE_ID>/):
        expression_text_preservado.csv
        expression_numerico.csv
        expression_analysis_ready.csv
        feature_map.csv
        expression_merge_ready.csv
        expression_merge_ready_zscore.csv
        sample_annotation.csv
        metadata_full.csv
        metadata_<filter>.csv

    Returns:
        Path to the output directory, or None on failure.
    """
    dataset_id = extract_dataset_id(path)
    out_dir = output_root / f"out_{dataset_id}"
    out_dir.mkdir(parents=True, exist_ok=True)

    log.info("")
    log.info("=" * 60)
    log.info(f"  DATASET: {dataset_id}")
    log.info(f"  File:    {path}")
    log.info(f"  Output:  {out_dir.absolute()}")
    log.info("=" * 60)

    # ── STEP 1: Metadata ─────────────────────────────────────────
    log.info("── Step 1: Metadata extraction ──")
    meta_df = parse_series_metadata_tabular(path)
    if meta_df.empty:
        log.error(f"Failed to parse metadata from {path}")
        return None

    platform_id, platform_name = detect_platform(meta_df)

    # ── STEP 2: Condition selection ──────────────────────────────
    log.info("── Step 2: Condition filtering ──")
    conditions, condition_cols = extract_conditions(meta_df)
    selected = select_conditions_cli(
        conditions, condition_filter, no_interactive,
    )
    meta_filtered = filter_samples_by_conditions(
        meta_df, selected, condition_cols,
    )

    # Save metadata
    meta_df.to_csv(out_dir / "metadata_full.csv", index=False)
    if selected:
        label = "_".join(
            re.sub(r"[^\w\s-]", "", c).strip().replace(" ", "_")
            for c in selected
        )
        meta_filtered.to_csv(
            out_dir / f"metadata_{label}.csv", index=False,
        )

    # ── STEP 3: Expression reading ───────────────────────────────
    log.info("── Step 3: Expression reading ──")
    expr_text, expr_num, gsm_cols = read_expression(path)

    # ── STEP 4: Cross-reference with filtered samples ────────────
    log.info("── Step 4: Cross-referencing samples ──")
    filtered_gsms: set = set()
    for col in meta_filtered.columns:
        vals = meta_filtered[col].astype(str)
        filtered_gsms.update(v for v in vals if v.startswith("GSM"))

    if selected and filtered_gsms:
        keep_cols = [c for c in gsm_cols if c in filtered_gsms]
        if not keep_cols:
            log.warning(
                "No matching GSM columns after filter; using ALL columns"
            )
            keep_cols = gsm_cols
    else:
        keep_cols = gsm_cols

    log.info(f"Samples retained: {len(keep_cols)}")

    expr_text_filt = expr_text[["Probe_ID"] + keep_cols].copy()
    expr_num_filt = expr_num[["Probe_ID"] + keep_cols].copy()

    # ── STEP 5: Scale inference ──────────────────────────────────
    log.info("── Step 5: Scale inference ──")
    scale = infer_scale(meta_df, platform_id, expr_num_filt)

    expr_ready = expr_num_filt.copy()
    if scale == "needs_log2":
        log.info("Applying log2(x + 1) transformation")
        for c in keep_cols:
            expr_ready[c] = np.log2(expr_ready[c].clip(lower=0) + 1)
    elif scale in ("already_log2", "already_processed"):
        log.info(f"Scale='{scale}' → no transformation applied")
    else:
        log.warning(
            f"Scale='{scale}' → keeping values as-is. Please verify manually."
        )

    # ── STEP 6: Probe ID harmonization ───────────────────────────
    log.info("── Step 6: Probe ID harmonization ──")
    feature_map = build_feature_map(expr_num_filt, platform_id)

    # Merge ready: canonical IDs, exclude ambiguous probes
    fmap_clean = feature_map[~feature_map["Probe_ID_Ambiguous"]].copy()

    expr_merge = expr_ready.merge(
        fmap_clean[["Probe_ID", "Probe_ID_Canonical"]],
        on="Probe_ID",
        how="inner",
    )
    expr_merge.drop(columns=["Probe_ID"], inplace=True)
    expr_merge.rename(
        columns={"Probe_ID_Canonical": "Probe_ID"}, inplace=True,
    )

    # Average duplicate canonical IDs
    if expr_merge["Probe_ID"].duplicated().any():
        n_dup = expr_merge["Probe_ID"].duplicated().sum()
        log.info(f"Averaging {n_dup} duplicate canonical Probe IDs")
        expr_merge = (
            expr_merge
            .groupby("Probe_ID", as_index=False)[keep_cols]
            .mean()
        )

    # Reorder columns
    expr_merge = expr_merge[["Probe_ID"] + keep_cols]

    # ── STEP 7: Z-score ──────────────────────────────────────────
    log.info("── Step 7: Z-score normalization ──")
    expr_zscore = zscore_by_probe(expr_merge)

    # ── STEP 8: Sample annotation ────────────────────────────────
    log.info("── Step 8: Sample annotation ──")
    sample_annot = build_sample_annotation(
        meta_filtered, keep_cols, dataset_id,
        platform_id, platform_name, class_map,
    )

    # ── SAVE ALL OUTPUTS ─────────────────────────────────────────
    log.info("── Saving outputs ──")
    expr_text_filt.to_csv(out_dir / "expression_text_preservado.csv", index=False)
    expr_num_filt.to_csv(out_dir / "expression_numerico.csv", index=False)
    expr_ready.to_csv(out_dir / "expression_analysis_ready.csv", index=False)
    feature_map.to_csv(out_dir / "feature_map.csv", index=False)
    expr_merge.to_csv(out_dir / "expression_merge_ready.csv", index=False)
    expr_zscore.to_csv(out_dir / "expression_merge_ready_zscore.csv", index=False)
    sample_annot.to_csv(out_dir / "sample_annotation.csv", index=False)

    log.info(f"✅ Dataset {dataset_id} — all outputs saved:")
    for fpath in sorted(out_dir.glob("*.csv")):
        size_kb = fpath.stat().st_size / 1024
        log.info(f"   • {fpath.name} ({size_kb:.1f} KB)")

    return out_dir


# =====================================================================
# 13. Cross-Dataset Merge
# =====================================================================

def merge_datasets(
    dataset_dirs: List[Path],
    output_dir: Path,
) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """
    Merge multiple datasets by shared (intersected) miRNA IDs.

    Reads BOTH expression_merge_ready.csv (raw log2) and
    expression_merge_ready_zscore.csv from each dataset directory,
    plus sample_annotation.csv.

    ComBat needs the RAW (non-z-scored) merged data to detect
    batch-mean differences. Z-scored data is also produced for
    reference but should NOT be used as ComBat input.

    Outputs:
        merged_expression_raw.csv       ← for ComBat input
        merged_expression_zscore.csv    ← for reference / fallback
        merged_sample_annotation.csv

    Returns:
        (merged_raw, merged_zscore, merged_annot)
    """
    log.info("")
    log.info("=" * 60)
    log.info("  MERGE: Combining datasets")
    log.info("=" * 60)

    all_raw: List[pd.DataFrame] = []
    all_zscore: List[pd.DataFrame] = []
    all_annot: List[pd.DataFrame] = []
    all_mirna_sets: List[set] = []

    for d in dataset_dirs:
        raw_path = d / "expression_merge_ready.csv"
        zscore_path = d / "expression_merge_ready_zscore.csv"
        annot_path = d / "sample_annotation.csv"

        if not raw_path.exists():
            log.warning(f"Missing {raw_path}; skipping {d.name}")
            continue
        if not zscore_path.exists():
            log.warning(f"Missing {zscore_path}; skipping {d.name}")
            continue
        if not annot_path.exists():
            log.warning(f"Missing {annot_path}; skipping {d.name}")
            continue

        raw = pd.read_csv(raw_path)
        zscore = pd.read_csv(zscore_path)
        annot = pd.read_csv(annot_path)

        mirnas = set(raw["Probe_ID"].unique())
        all_mirna_sets.append(mirnas)
        all_raw.append(raw)
        all_zscore.append(zscore)
        all_annot.append(annot)

        log.info(
            f"  {d.name}: {raw.shape[0]} miRNAs, "
            f"{raw.shape[1] - 1} samples"
        )

    if len(all_raw) < 2:
        log.warning("Need >= 2 datasets to merge")
        if all_raw:
            return all_raw[0], all_zscore[0], all_annot[0]
        return pd.DataFrame(), pd.DataFrame(), pd.DataFrame()

    # Intersect miRNA IDs
    common: set = all_mirna_sets[0]
    for s in all_mirna_sets[1:]:
        common = common.intersection(s)

    log.info(f"Common miRNAs across all datasets: {len(common)}")

    if not common:
        log.error(
            "No common miRNAs found! "
            "Verify probe ID harmonization."
        )
        return pd.DataFrame(), pd.DataFrame(), pd.DataFrame()

    # Subset each to common miRNAs and concatenate columns
    def _merge_list(dfs: List[pd.DataFrame]) -> pd.DataFrame:
        parts = []
        for df in dfs:
            sub = df[df["Probe_ID"].isin(common)].copy()
            sub = sub.set_index("Probe_ID")
            parts.append(sub)
        return pd.concat(parts, axis=1).reset_index()

    merged_raw = _merge_list(all_raw)
    merged_zscore = _merge_list(all_zscore)
    merged_annot = pd.concat(all_annot, ignore_index=True)

    # Save
    output_dir.mkdir(parents=True, exist_ok=True)
    merged_raw.to_csv(
        output_dir / "merged_expression_raw.csv", index=False,
    )
    merged_zscore.to_csv(
        output_dir / "merged_expression_zscore.csv", index=False,
    )
    merged_annot.to_csv(
        output_dir / "merged_sample_annotation.csv", index=False,
    )

    log.info(
        f"Merged: {merged_raw.shape[0]} miRNAs x "
        f"{merged_raw.shape[1] - 1} samples"
    )
    log.info(
        f"Samples by dataset: "
        f"{merged_annot['dataset_id'].value_counts().to_dict()}"
    )
    log.info(
        f"Samples by class: "
        f"{merged_annot['class_label'].value_counts().to_dict()}"
    )

    return merged_raw, merged_zscore, merged_annot


# =====================================================================
# 14. ComBat Batch Correction
# =====================================================================

def _try_install_neurocombat() -> bool:
    """Attempt to install neuroCombat via pip."""
    log.info("Attempting: pip install neuroCombat ...")
    try:
        result = subprocess.run(
            [sys.executable, "-m", "pip", "install", "neuroCombat"],
            capture_output=True, text=True, timeout=120,
        )
        if result.returncode == 0:
            log.info("✅ neuroCombat installed successfully")
            return True
        log.warning(f"pip install failed:\n{result.stderr[:500]}")
        return False
    except Exception as exc:
        log.warning(f"Install failed: {exc}")
        return False


def apply_combat(
    expr_df: pd.DataFrame,
    sample_annot: pd.DataFrame,
    batch_col: str = "batch",
    class_col: str = "class_label",
) -> Optional[pd.DataFrame]:
    """
    Apply ComBat batch correction preserving biological signal.

    The class_label column is used as a biological covariate so that
    disease-related variation is NOT removed.

    Tries libraries in order:
        1. neuroCombat  (auto-installs if missing)
        2. inmoose.pycombat.pycombat_norm
        3. Returns None with instructions

    Args:
        expr_df:      Probe_ID + GSM columns (features × samples)
        sample_annot: sample_id, batch, class_label, …
        batch_col:    column name for batch (dataset/platform)
        class_col:    column name for biological class

    Returns:
        Corrected DataFrame (same shape as input), or None.
    """
    gsm_cols = [c for c in expr_df.columns if c.startswith("GSM")]

    # Align samples between expression and annotation
    annot_ids = set(sample_annot["sample_id"])
    common_samples = [s for s in gsm_cols if s in annot_ids]

    if len(common_samples) < len(gsm_cols):
        log.warning(
            f"Annotation has {len(common_samples)}/{len(gsm_cols)} "
            f"matching samples"
        )

    annot_aligned = (
        sample_annot
        .set_index("sample_id")
        .loc[common_samples]
        .reset_index()
    )

    # Expression matrix: features × samples (numpy)
    expr_matrix = (
        expr_df
        .set_index("Probe_ID")[common_samples]
        .values
        .astype(float)
    )

    # Remove rows that are entirely NaN
    valid_mask = ~np.all(np.isnan(expr_matrix), axis=1)
    expr_clean = expr_matrix[valid_mask].copy()
    probe_ids = expr_df["Probe_ID"].values[valid_mask]

    # Impute remaining NaNs with row mean
    for i in range(expr_clean.shape[0]):
        row = expr_clean[i]
        nans = np.isnan(row)
        if nans.any():
            row[nans] = np.nanmean(row)
            expr_clean[i] = row

    # Covariates DataFrame
    covars = pd.DataFrame({
        batch_col: annot_aligned[batch_col].values,
        class_col: annot_aligned[class_col].values,
    })

    log.info(
        f"ComBat input: {expr_clean.shape[0]} features × "
        f"{expr_clean.shape[1]} samples"
    )
    log.info(f"  Batches: {covars[batch_col].value_counts().to_dict()}")
    log.info(f"  Classes: {covars[class_col].value_counts().to_dict()}")

    corrected = None

    # ── Try 1: neuroCombat ──
    try:
        from neuroCombat import neuroCombat

        log.info("Using neuroCombat for batch correction")
        result = neuroCombat(
            dat=expr_clean,
            covars=covars,
            batch_col=batch_col,
            categorical_cols=[class_col],
        )
        corrected = result["data"]
        log.info("✅ neuroCombat completed successfully")

    except ImportError:
        log.info("neuroCombat not found — attempting auto-install...")
        if _try_install_neurocombat():
            try:
                from neuroCombat import neuroCombat

                result = neuroCombat(
                    dat=expr_clean,
                    covars=covars,
                    batch_col=batch_col,
                    categorical_cols=[class_col],
                )
                corrected = result["data"]
                log.info("✅ neuroCombat completed (after install)")
            except Exception as exc:
                log.error(f"neuroCombat failed after install: {exc}")
        else:
            log.info("neuroCombat auto-install failed; trying inmoose…")

    except Exception as exc:
        log.error(f"neuroCombat error: {exc}")

    # ── Try 2: inmoose ──
    if corrected is None:
        try:
            from inmoose.pycombat import pycombat_norm

            log.info("Using inmoose pycombat_norm for batch correction")

            # Build design matrix for biological covariate
            covar_mod = pd.get_dummies(
                covars[[class_col]], drop_first=True,
            ).astype(float)

            expr_df_in = pd.DataFrame(
                expr_clean, index=probe_ids, columns=common_samples,
            )

            corrected_df = pycombat_norm(
                counts=expr_df_in,
                batch=covars[batch_col].values,
                covar_mod=covar_mod,
            )
            corrected = corrected_df.values
            log.info("✅ inmoose pycombat_norm completed successfully")

        except ImportError:
            log.error(
                "❌ No ComBat library available!\n"
                "   Install one of:\n"
                "     pip install neuroCombat\n"
                "     pip install inmoose\n"
            )
            return None
        except Exception as exc:
            log.error(f"inmoose error: {exc}")
            return None

    # ── Rebuild DataFrame ──
    out_df = pd.DataFrame(corrected, columns=common_samples)
    out_df.insert(0, "Probe_ID", probe_ids)

    log.info(
        f"ComBat output: {out_df.shape[0]} features × "
        f"{out_df.shape[1] - 1} samples"
    )
    return out_df


# =====================================================================
# 15. Purity Metrics (PurityB & PurityD)
# =====================================================================

def calculate_purity(
    cluster_labels: np.ndarray,
    true_labels: np.ndarray,
) -> float:
    """
    Calculate cluster purity.

        Purity = (1/N) * Σ_k  max_j  |w_k ∩ c_j|

    where:
        N   = total number of samples
        w_k = set of samples assigned to cluster k
        c_j = set of samples belonging to true class j
    """
    n = len(cluster_labels)
    if n == 0:
        return 0.0

    cluster_ids = np.unique(cluster_labels)
    true_ids = np.unique(true_labels)

    total = 0
    for k in cluster_ids:
        cluster_mask = cluster_labels == k
        max_overlap = 0
        for j in true_ids:
            class_mask = true_labels == j
            overlap = int(np.sum(cluster_mask & class_mask))
            if overlap > max_overlap:
                max_overlap = overlap
        total += max_overlap

    return total / n


def compute_purity_metrics(
    expr_before: pd.DataFrame,
    expr_after: Optional[pd.DataFrame],
    sample_annot: pd.DataFrame,
    batch_col: str = "batch",
    class_col: str = "class_label",
) -> pd.DataFrame:
    """
    Compute PurityB and PurityD before and after ComBat.

    PurityB: K-Means with k = n_batches, true labels = batch
             → should DECREASE after ComBat (batches become mixed)
    PurityD: K-Means with k = n_classes, true labels = class_label
             → should STAY HIGH or INCREASE after ComBat

    Returns:
        DataFrame with one row: PurityB_before, PurityB_after,
        PurityD_before, PurityD_after
    """
    gsm_cols_before = [
        c for c in expr_before.columns if c.startswith("GSM")
    ]
    common = [
        s for s in gsm_cols_before
        if s in sample_annot["sample_id"].values
    ]
    annot = sample_annot.set_index("sample_id").loc[common]

    batches = annot[batch_col].values
    classes = annot[class_col].values

    n_batches = len(np.unique(batches))
    n_classes = len(np.unique(classes))

    le_batch = LabelEncoder()
    batch_enc = le_batch.fit_transform(batches)
    le_class = LabelEncoder()
    class_enc = le_class.fit_transform(classes)

    results: Dict[str, float] = {}

    # ── Before ComBat ──
    X_before = (
        expr_before.set_index("Probe_ID")[common].values.T
    )  # samples × features
    X_before = np.nan_to_num(X_before, nan=0.0)

    km_b = KMeans(n_clusters=n_batches, random_state=42, n_init=10)
    results["PurityB_before"] = calculate_purity(
        km_b.fit_predict(X_before), batch_enc,
    )

    km_d = KMeans(n_clusters=n_classes, random_state=42, n_init=10)
    results["PurityD_before"] = calculate_purity(
        km_d.fit_predict(X_before), class_enc,
    )

    # ── After ComBat ──
    if expr_after is not None:
        gsm_after = [
            c for c in expr_after.columns if c.startswith("GSM")
        ]
        common_after = [s for s in gsm_after if s in common]

        X_after = (
            expr_after.set_index("Probe_ID")[common_after].values.T
        )
        X_after = np.nan_to_num(X_after, nan=0.0)

        annot_after = sample_annot.set_index("sample_id").loc[common_after]
        batch_after = le_batch.transform(annot_after[batch_col].values)
        class_after = le_class.transform(annot_after[class_col].values)

        km_b2 = KMeans(n_clusters=n_batches, random_state=42, n_init=10)
        results["PurityB_after"] = calculate_purity(
            km_b2.fit_predict(X_after), batch_after,
        )

        km_d2 = KMeans(n_clusters=n_classes, random_state=42, n_init=10)
        results["PurityD_after"] = calculate_purity(
            km_d2.fit_predict(X_after), class_after,
        )
    else:
        results["PurityB_after"] = np.nan
        results["PurityD_after"] = np.nan

    # Print summary
    log.info("")
    log.info("=" * 55)
    log.info("  PURITY METRICS")
    log.info("=" * 55)
    for key, val in results.items():
        log.info(f"  {key:<20s}: {val:.4f}")
    log.info("-" * 55)
    log.info(
        "  ✓ PurityB should DECREASE after ComBat "
        "(batch effect removed)"
    )
    log.info(
        "  ✓ PurityD should STAY HIGH or INCREASE "
        "(biological signal preserved)"
    )
    log.info("=" * 55)

    return pd.DataFrame([results])


# =====================================================================
# 16. Visualization
# =====================================================================

def _pca_scatter(
    expr_df: pd.DataFrame,
    sample_annot: pd.DataFrame,
    color_col: str,
    title: str,
    save_path: Path,
) -> None:
    """Generate and save a PCA scatter plot."""
    gsm_cols = [c for c in expr_df.columns if c.startswith("GSM")]
    common = [
        s for s in gsm_cols
        if s in sample_annot["sample_id"].values
    ]

    X = expr_df.set_index("Probe_ID")[common].values.T  # samples × features
    X = np.nan_to_num(X, nan=0.0)

    pca = PCA(n_components=2, random_state=42)
    pc = pca.fit_transform(X)

    annot = sample_annot.set_index("sample_id").loc[common]
    labels = annot[color_col].values
    unique_labels = sorted(set(labels))

    fig, ax = plt.subplots(figsize=(10, 8))

    if HAS_SEABORN:
        palette = sns.color_palette("husl", len(unique_labels))
    else:
        cmap = plt.cm.get_cmap("tab10", max(len(unique_labels), 1))
        palette = [cmap(i) for i in range(len(unique_labels))]

    for i, label in enumerate(unique_labels):
        mask = labels == label
        ax.scatter(
            pc[mask, 0], pc[mask, 1],
            label=label,
            alpha=0.7,
            s=50,
            color=palette[i % len(palette)],
            edgecolors="white",
            linewidth=0.5,
        )

    ev = pca.explained_variance_ratio_
    ax.set_xlabel(f"PC1 ({ev[0] * 100:.1f}%)")
    ax.set_ylabel(f"PC2 ({ev[1] * 100:.1f}%)")
    ax.set_title(title, fontsize=13)
    ax.legend(
        bbox_to_anchor=(1.05, 1), loc="upper left",
        fontsize=9, framealpha=0.9,
    )

    fig.tight_layout()
    fig.savefig(save_path, dpi=150, bbox_inches="tight")
    plt.close(fig)
    log.info(f"🖼️  Plot saved: {save_path.name}")


def _umap_scatter(
    expr_df: pd.DataFrame,
    sample_annot: pd.DataFrame,
    color_col: str,
    title: str,
    save_path: Path,
) -> None:
    """Generate and save a UMAP scatter plot (requires umap-learn)."""
    if not HAS_UMAP:
        return

    gsm_cols = [c for c in expr_df.columns if c.startswith("GSM")]
    common = [
        s for s in gsm_cols
        if s in sample_annot["sample_id"].values
    ]

    X = expr_df.set_index("Probe_ID")[common].values.T
    X = np.nan_to_num(X, nan=0.0)

    reducer = umap_lib.UMAP(n_components=2, random_state=42)
    embedding = reducer.fit_transform(X)

    annot = sample_annot.set_index("sample_id").loc[common]
    labels = annot[color_col].values
    unique_labels = sorted(set(labels))

    fig, ax = plt.subplots(figsize=(10, 8))

    if HAS_SEABORN:
        palette = sns.color_palette("husl", len(unique_labels))
    else:
        cmap = plt.cm.get_cmap("tab10", max(len(unique_labels), 1))
        palette = [cmap(i) for i in range(len(unique_labels))]

    for i, label in enumerate(unique_labels):
        mask = labels == label
        ax.scatter(
            embedding[mask, 0], embedding[mask, 1],
            label=label,
            alpha=0.7,
            s=50,
            color=palette[i % len(palette)],
            edgecolors="white",
            linewidth=0.5,
        )

    ax.set_xlabel("UMAP 1")
    ax.set_ylabel("UMAP 2")
    ax.set_title(title, fontsize=13)
    ax.legend(
        bbox_to_anchor=(1.05, 1), loc="upper left",
        fontsize=9, framealpha=0.9,
    )

    fig.tight_layout()
    fig.savefig(save_path, dpi=150, bbox_inches="tight")
    plt.close(fig)
    log.info(f"🖼️  UMAP saved: {save_path.name}")


def generate_all_plots(
    expr_before: pd.DataFrame,
    expr_after: Optional[pd.DataFrame],
    sample_annot: pd.DataFrame,
    output_dir: Path,
) -> None:
    """Generate PCA (and optional UMAP) plots before/after ComBat."""
    log.info("── Generating visualizations ──")

    # Before ComBat
    _pca_scatter(
        expr_before, sample_annot, "batch",
        "PCA — Before ComBat (colored by Batch/Platform)",
        output_dir / "pca_before_batch.png",
    )
    _pca_scatter(
        expr_before, sample_annot, "class_label",
        "PCA — Before ComBat (colored by Class)",
        output_dir / "pca_before_class.png",
    )

    # After ComBat
    if expr_after is not None:
        _pca_scatter(
            expr_after, sample_annot, "batch",
            "PCA — After ComBat (colored by Batch/Platform)",
            output_dir / "pca_after_batch.png",
        )
        _pca_scatter(
            expr_after, sample_annot, "class_label",
            "PCA — After ComBat (colored by Class)",
            output_dir / "pca_after_class.png",
        )

    # Optional UMAP
    if HAS_UMAP:
        log.info("UMAP available — generating UMAP plots")
        _umap_scatter(
            expr_before, sample_annot, "batch",
            "UMAP — Before ComBat (Batch)",
            output_dir / "umap_before_batch.png",
        )
        _umap_scatter(
            expr_before, sample_annot, "class_label",
            "UMAP — Before ComBat (Class)",
            output_dir / "umap_before_class.png",
        )
        if expr_after is not None:
            _umap_scatter(
                expr_after, sample_annot, "batch",
                "UMAP — After ComBat (Batch)",
                output_dir / "umap_after_batch.png",
            )
            _umap_scatter(
                expr_after, sample_annot, "class_label",
                "UMAP — After ComBat (Class)",
                output_dir / "umap_after_class.png",
            )
    else:
        log.info("UMAP not available (install umap-learn for UMAP plots)")


# =====================================================================
# 17. Main
# =====================================================================

def main() -> None:
    """Main pipeline entry point."""
    args = build_cli()

    print()
    print("=" * 60)
    print("  GEO miRNA Cross-Platform Integration Pipeline")
    print("  For PDAC (Pancreatic Cancer) Research")
    print("=" * 60)

    # ── Interactive mode: if no files provided, open file picker ──
    if not args.files:
        print()
        print("  Nenhum arquivo informado via linha de comando.")
        print("  Vamos selecionar os arquivos de forma interativa!")
        print()
        print("-" * 50)
        print("  Como deseja selecionar os arquivos?")
        print("-" * 50)
        print("  [1] Abrir seletor de arquivos (explorador)")
        print("  [2] Digitar os caminhos manualmente")
        print("-" * 50)

        file_choice = ""
        while file_choice not in ("1", "2"):
            try:
                file_choice = input(
                    "\n  Escolha (1/2) [padrao: 1]: "
                ).strip() or "1"
            except (EOFError, KeyboardInterrupt):
                print("\n  Cancelado pelo usuario.")
                sys.exit(0)

        if file_choice == "1":
            selected_files = interactive_file_picker()
        else:
            print("\n  Digite os caminhos dos arquivos (um por linha).")
            print("  Quando terminar, digite uma linha vazia e pressione Enter.")
            selected_files = []
            while True:
                try:
                    line = input("  Arquivo: ").strip().strip('"')
                    if not line:
                        break
                    selected_files.append(line)
                except (EOFError, KeyboardInterrupt):
                    break

        if not selected_files:
            print("\n  Nenhum arquivo selecionado. Encerrando.")
            sys.exit(0)

        args.files = selected_files

        # Show selected files
        print("\n" + "-" * 50)
        print(f"  {len(args.files)} arquivo(s) selecionado(s):")
        print("-" * 50)
        for i, f in enumerate(args.files, 1):
            print(f"  {i}. {f}")

        # Ask for output directory
        out_choice = interactive_output_picker()
        args.output_root = out_choice

        print("\n" + "=" * 60)
        print("  Configuracao:")
        print(f"  Arquivos:    {len(args.files)}")
        print(f"  Saida:       {Path(args.output_root).absolute()}")
        print("=" * 60)

        try:
            confirm = input("\n  Iniciar pipeline? (S/n): ").strip().lower()
            if confirm in ("n", "nao", "no"):
                print("  Cancelado pelo usuario.")
                sys.exit(0)
        except (EOFError, KeyboardInterrupt):
            print("\n  Cancelado pelo usuario.")
            sys.exit(0)

    output_root = Path(args.output_root)
    output_root.mkdir(parents=True, exist_ok=True)

    # Parse class map from CLI
    class_map: Optional[Dict[str, str]] = None
    if args.class_map:
        class_map = {}
        for item in args.class_map:
            if "=" in item:
                key, val = item.split("=", 1)
                class_map[key.strip()] = val.strip()
        log.info(f"Class mapping: {class_map}")

    # ────────────────────────────────────────────────────────────
    # Phase 1: Process each dataset individually
    # ────────────────────────────────────────────────────────────
    dataset_dirs: List[Path] = []

    for filepath in args.files:
        if not Path(filepath).exists():
            log.error(f"File not found: {filepath}")
            continue

        result_dir = process_single_dataset(
            path=filepath,
            output_root=output_root,
            no_interactive=args.no_interactive,
            condition_filter=args.condition_filter,
            class_map=class_map,
        )
        if result_dir is not None:
            dataset_dirs.append(result_dir)

    if not dataset_dirs:
        log.error("No datasets processed successfully. Exiting.")
        sys.exit(1)

    # ────────────────────────────────────────────────────────────
    # Phase 2: Merge + ComBat + Validation (if ≥ 2 datasets)
    # ────────────────────────────────────────────────────────────
    if len(dataset_dirs) >= 2:
        # Merge
        merged_raw, merged_zscore, merged_annot = merge_datasets(
            dataset_dirs, output_root,
        )

        if merged_raw.empty:
            log.error("Merge failed — no output generated.")
            sys.exit(1)

        # ComBat — applied to RAW (non-z-scored) merged data
        # so that batch-mean differences are visible and correctable
        expr_combat: Optional[pd.DataFrame] = None
        if not args.no_combat and not args.zscore_only:
            log.info("")
            log.info("── ComBat Batch Correction ──")
            expr_combat = apply_combat(merged_raw, merged_annot)

            if expr_combat is not None:
                expr_combat.to_csv(
                    output_root / "merged_expression_combat.csv",
                    index=False,
                )
                # Also produce a z-scored version of ComBat output
                expr_combat_zscore = zscore_by_probe(expr_combat)
                expr_combat_zscore.to_csv(
                    output_root / "merged_expression_combat_zscore.csv",
                    index=False,
                )
                log.info(
                    "Saved: merged_expression_combat.csv, "
                    "merged_expression_combat_zscore.csv"
                )

        # Purity validation — uses the RAW data (before/after ComBat)
        log.info("")
        log.info("── Purity Validation ──")
        purity_df = compute_purity_metrics(
            merged_raw, expr_combat, merged_annot,
        )
        purity_df.to_csv(
            output_root / "purity_metrics.csv", index=False,
        )

        # Plots — show the RAW data before and ComBat output after
        if not args.no_plots:
            generate_all_plots(
                merged_raw, expr_combat, merged_annot, output_root,
            )

    else:
        log.info(
            "Single dataset processed — "
            "skipping merge / ComBat / purity steps."
        )

    # ────────────────────────────────────────────────────────────
    # Final Summary
    # ────────────────────────────────────────────────────────────
    print("")
    print("=" * 60)
    print("  🎉 Pipeline completed successfully!")
    print("=" * 60)
    print(f"  📁 Output root: {output_root.absolute()}")
    print(f"  📊 Datasets processed: {len(dataset_dirs)}")
    for d in dataset_dirs:
        n_csv = len(list(d.glob("*.csv")))
        print(f"     └─ {d.name}/ ({n_csv} CSV files)")

    if len(dataset_dirs) >= 2:
        print("  📊 Merged outputs:")
        for fpath in sorted(output_root.glob("merged_*")):
            size_kb = fpath.stat().st_size / 1024
            print(f"     • {fpath.name} ({size_kb:.1f} KB)")
        for fpath in sorted(output_root.glob("purity_*")):
            print(f"     • {fpath.name}")
        for fpath in sorted(output_root.glob("*.png")):
            print(f"     🖼️ {fpath.name}")

    print("=" * 60)


if __name__ == "__main__":
    main()
